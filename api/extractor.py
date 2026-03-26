import os
import json
import base64
import re
import pandas as pd
import fitz  # PyMuPDF
from litellm import completion, completion_cost
from concurrent.futures import ThreadPoolExecutor, as_completed


# ---------------------------------------------------------------------------
# PDF Utilities
# ---------------------------------------------------------------------------

def extract_images_from_pdf(pdf_path: str) -> list:
    """Extract all pages as base64 JPEG images for vision-based OCR."""
    images = []
    try:
        doc = fitz.open(pdf_path)
        for page in doc:
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom for clarity
            img_bytes = pix.tobytes("jpeg")
            b64 = base64.b64encode(img_bytes).decode("utf-8")
            images.append(b64)
        return images
    except Exception as e:
        print(f"Error converting PDF to images {pdf_path}: {e}")
        return []


def extract_text_from_pdf(pdf_path: str) -> str:
    """Extract text from PDF preserving reading order (top-to-bottom, left-to-right)."""
    try:
        doc = fitz.open(pdf_path)
        text = ""
        for page in doc:
            blocks = page.get_text("blocks")
            blocks.sort(key=lambda b: (b[1], b[0]))
            for b in blocks:
                if b[4].strip():
                    text += b[4] + "\n"
        return text
    except Exception as e:
        print(f"Error reading PDF {pdf_path}: {e}")
        return ""


# ---------------------------------------------------------------------------
# Regex — DATES ONLY
# ---------------------------------------------------------------------------

def extract_date_hints(text: str) -> dict:
    """
    Use regex ONLY for date extraction as hints passed to the LLM.
    The LLM remains the sole authority on which date is correct and
    is used as the fallback when regex finds nothing.
    """
    date_patterns = [
        # MM/DD/YYYY  or  MM-DD-YYYY
        r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b',
        # YYYY-MM-DD
        r'\b\d{4}[/-]\d{1,2}[/-]\d{1,2}\b',
        # Month DD, YYYY  (e.g. January 15, 2024)
        r'\b(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|'
        r'Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|'
        r'Dec(?:ember)?)\s+\d{1,2},?\s+\d{4}\b',
        # DD Month YYYY  (e.g. 15 January 2024)
        r'\b\d{1,2}\s+(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|'
        r'Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|'
        r'Nov(?:ember)?|Dec(?:ember)?)\s+\d{4}\b',
    ]

    found: list = []
    for pattern in date_patterns:
        found.extend(re.findall(pattern, text, re.IGNORECASE))

    # Deduplicate while preserving order
    seen: set = set()
    unique: list = []
    for d in found:
        if d not in seen:
            seen.add(d)
            unique.append(d)

    return {"date_hints": unique[:15]} if unique else {}


# ---------------------------------------------------------------------------
# Smart chunking — NO content filtering; every byte of text is preserved
# ---------------------------------------------------------------------------

def chunk_text_smartly(text: str, max_chars: int = 80_000, overlap: int = 3_000) -> list:
    """
    Split text into overlapping chunks at natural paragraph/line boundaries.
    Overlap prevents any record that straddles a chunk boundary from being lost.

    For documents <= max_chars the original text is returned as a single-item list.
    """
    if len(text) <= max_chars:
        return [text]

    chunks = []
    start = 0

    while start < len(text):
        end = start + max_chars

        if end >= len(text):
            chunks.append(text[start:])
            break

        # Prefer splitting at a blank-line boundary, then a newline, then a space
        split_pos = text.rfind('\n\n', start + max_chars // 2, end)
        if split_pos == -1:
            split_pos = text.rfind('\n', start + max_chars // 2, end)
        if split_pos == -1:
            split_pos = text.rfind(' ', start + max_chars // 2, end)
        if split_pos == -1:
            split_pos = end

        chunks.append(text[start:split_pos])
        # Roll back by `overlap` so boundary records are caught by both chunks
        start = max(split_pos - overlap, start + 1)

    return chunks


# ---------------------------------------------------------------------------
# Learning rules
# ---------------------------------------------------------------------------

def get_learning_rules() -> str:
    rules_path = "/tmp/rules.json"
    if not os.path.exists(rules_path):
        return ""
    try:
        with open(rules_path, "r") as f:
            rules = json.load(f)
        if not rules:
            return ""
        rule_texts = "\n".join([f"- {r['rule']}" for r in rules])
        return f"\n\nCRITICAL AI LEARNING RULES (apply to ALL extraction):\n{rule_texts}"
    except Exception as e:
        print(f"Error loading rules: {e}")
        return ""


# ---------------------------------------------------------------------------
# Deduplication (for overlapping chunks)
# ---------------------------------------------------------------------------

def deduplicate_records(records: list, col_names: list) -> list:
    """
    Remove duplicate records that may appear in overlapping chunk windows.
    A record is considered a duplicate when the first 3 non-N/A key fields match.
    """
    seen_keys: set = set()
    unique: list = []

    for record in records:
        key_parts = []
        for field in col_names:
            val = str(record.get(field, '')).strip().lower()
            if val and val not in ('n/a', '', 'none', 'null'):
                key_parts.append(val)
            if len(key_parts) == 3:
                break

        if key_parts:
            record_key = '|'.join(key_parts)
            if record_key not in seen_keys:
                seen_keys.add(record_key)
                unique.append(record)
        else:
            unique.append(record)

    return unique


# ---------------------------------------------------------------------------
# Core LLM call (text or vision)
# ---------------------------------------------------------------------------

def call_llm_for_extraction(
    prompt_instructions: str,
    text_chunk: str,
    api_key: str,
    model: str,
    images=None,
) -> tuple:
    """
    Execute one LLM extraction call.
    Returns (records: list[dict], tokens: int, cost: float).
    """
    tokens = 0
    cost = 0.0

    try:
        if images:
            content_array = [
                {"type": "text", "text": f"{prompt_instructions}\n\nDOCUMENT TEXT:\n{text_chunk}"}
            ]
            for b64_img in images:
                content_array.append({
                    "type": "image_url",
                    "image_url": {"url": f"data:image/jpeg;base64,{b64_img}"},
                })
            messages = [{"role": "user", "content": content_array}]
        else:
            full_prompt = f"{prompt_instructions}\n\nDOCUMENT TEXT:\n{text_chunk}"
            messages = [{"role": "user", "content": full_prompt}]

        response = completion(
            model=model,
            messages=messages,
            api_key=api_key,
            temperature=0.1,
            response_format={"type": "json_object"},
        )

        if hasattr(response, "usage") and response.usage:
            tokens = response.usage.total_tokens

        try:
            cost = completion_cost(completion_response=response)
        except Exception:
            pass

        content = response.choices[0].message.content.strip()
        records = _parse_llm_json(content)
        return records, tokens, cost

    except Exception as e:
        print(f"LLM call error: {e}")
        return [], 0, 0.0


# ---------------------------------------------------------------------------
# Fallback targeted re-extraction for fields that are universally N/A
# ---------------------------------------------------------------------------

def _fallback_targeted_extraction(
    records: list,
    col_names: list,
    columns: list,
    text: str,
    images: list,
    api_key: str,
    model: str,
) -> tuple:
    """
    For every field that came back as N/A across ALL records, fire a targeted
    LLM re-extraction sweep.  Only fields that are genuinely absent in every row
    trigger this fallback so cost remains minimal.

    Returns (updated_records, extra_tokens, extra_cost).
    """
    if not records or (not text.strip() and not images):
        return records, 0, 0.0

    total_tokens = 0
    total_cost = 0.0

    skip = {'Source File'}
    missing_fields = [
        col for col in col_names
        if col not in skip
        and all(
            str(r.get(col, 'N/A')).strip().lower() in ('n/a', '', 'none', 'null')
            for r in records
        )
    ]

    if not missing_fields:
        return records, 0, 0.0

    print(f"  Fallback sweep for: {missing_fields}")

    missing_col_defs = [
        c for c in columns
        if (c.get('name', c) if isinstance(c, dict) else c) in missing_fields
    ]

    fallback_prompt = (
        "You are a precise data extraction assistant.\n"
        "The following fields were NOT found in a previous extraction pass.\n"
        "Search the document CAREFULLY and return ONLY these fields.\n\n"
        f"Fields needed:\n{json.dumps(missing_fields)}\n\n"
        f"Field definitions:\n{json.dumps(missing_col_defs, indent=2)}\n\n"
        "Return a JSON object with key \"values\" containing an object with these fields as keys.\n"
        "Use \"N/A\" only if truly absent. Return ONLY valid JSON."
    )

    # Use first 40k chars — sufficient for any header/global data
    sample_text = text[:40_000] if text else ""
    sample_images = images[:2] if images else None

    fallback_records, tokens, cost = call_llm_for_extraction(
        fallback_prompt, sample_text, api_key, model, sample_images
    )
    total_tokens += tokens
    total_cost += cost

    if fallback_records:
        raw = fallback_records[0] if isinstance(fallback_records[0], dict) else {}
        fallback_values = raw.get("values", raw)

        for field in missing_fields:
            found_val = fallback_values.get(field, "N/A")
            if str(found_val).strip().lower() not in ('n/a', '', 'none', 'null'):
                for record in records:
                    record[field] = found_val

    return records, total_tokens, total_cost


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_col_names(columns: list) -> list:
    return [col.get('name', col) if isinstance(col, dict) else col for col in columns]


def _parse_llm_json(content: str) -> list:
    """Parse LLM JSON output into a list of dicts."""
    try:
        parsed = json.loads(content)
        if isinstance(parsed, dict) and "records" in parsed:
            inner = parsed["records"]
            return inner if isinstance(inner, list) else [inner]
        if isinstance(parsed, list):
            return parsed
        if isinstance(parsed, dict):
            if "values" in parsed and isinstance(parsed["values"], dict):
                return [parsed["values"]]
            return [parsed]
        return []
    except Exception:
        return []


def _build_prompt(col_names: list, columns: list, learning_rules: str, date_hints: dict) -> str:
    """Construct the extraction system prompt for a document."""
    prompt = (
        "You are an advanced data extraction assistant. "
        "You process medical documents (IDR determinations, EOBs, records, etc.)."
        f"{learning_rules}\n\n"
        "Extract ALL records from the document below. "
        "Return a JSON object with one key \"records\" whose value is an array of objects.\n\n"
        f"Each object must use EXACTLY these keys:\n{json.dumps(col_names)}\n\n"
        f"Extraction logic per field:\n{json.dumps(columns, indent=2)}\n\n"
        "RULES:\n"
        "1. Extract EVERY record (e.g. multiple Determination blocks / line items) — do not stop after the first.\n"
        "2. DATE RULE: Identify the single global document date (letter date, determination date, or "
        "\"period begins on\" date). Copy this EXACT value into EVERY row's Date field. "
        "NO row may have \"N/A\" for Date if any date exists anywhere in the document.\n"
        "3. DETERMINATION NUMBER RULE: If a single determination, use \"1\". "
        "If multiple line items, number them \"1\", \"2\", \"3\", etc. NEVER use N/A here.\n"
        "4. PREVAILING PARTY RULE: For Initiating / Non-Initiating columns output "
        "only \"X\" (winner) or \"N/A\" (loser). Never output numbers.\n"
        "5. Do NOT create isolated header-only rows; attach all global data to each record.\n"
        "6. Keys must exactly match the field names listed above. Use \"N/A\" for any missing value.\n"
        "7. Return ONLY valid JSON — no markdown, no backticks, no commentary."
    )

    if date_hints:
        prompt += (
            f"\n\nDATE HINTS (regex-extracted from document — verify and use the correct one):\n"
            f"{json.dumps(date_hints, indent=2)}"
        )

    return prompt


# ---------------------------------------------------------------------------
# Single-PDF processor
# ---------------------------------------------------------------------------

def process_single_pdf(pdf_path: str, columns: list, api_key: str, learning_rules: str) -> dict:
    """
    Extract all records from one PDF.
    Returns {'rows': [...], 'total_tokens': int, 'total_cost': float}.
    """
    filename = os.path.basename(pdf_path)
    text = extract_text_from_pdf(pdf_path)
    col_names = _get_col_names(columns)

    # Detect scanned / image-only PDF
    is_scanned = False
    page_count = 1
    try:
        doc = fitz.open(pdf_path)
        page_count = len(doc)
        if len(text.strip()) < page_count * 50:
            is_scanned = True
        doc.close()
    except Exception:
        pass

    model = "gpt-4o-mini"

    images: list = []
    if is_scanned:
        images = extract_images_from_pdf(pdf_path)

    if not text.strip() and not images:
        row = {c: f"Empty/Unreadable PDF: {filename}" for c in col_names}
        row['Source File'] = filename
        return {'rows': [row], 'total_tokens': 0, 'total_cost': 0.0}

    if not api_key:
        row = {c: "Error: Missing OPENAI_API_KEY" for c in col_names}
        row['Source File'] = filename
        return {'rows': [row], 'total_tokens': 0, 'total_cost': 0.0}

    # Regex ONLY for dates — passed as hints, LLM makes the final call
    date_hints = extract_date_hints(text) if (not is_scanned and text) else {}

    prompt_instructions = _build_prompt(col_names, columns, learning_rules, date_hints)

    total_tokens = 0
    total_cost = 0.0
    all_records: list = []

    if is_scanned and images:
        # Vision path: batch pages to stay within token/time limits
        PAGES_PER_BATCH = 8
        for i in range(0, len(images), PAGES_PER_BATCH):
            batch = images[i: i + PAGES_PER_BATCH]
            label = f"[Pages {i + 1} to {min(i + PAGES_PER_BATCH, len(images))} of {len(images)}]"
            records, tokens, cost = call_llm_for_extraction(
                prompt_instructions, label, api_key, model, batch
            )
            all_records.extend(records)
            total_tokens += tokens
            total_cost += cost
    else:
        # Text path: smart chunking — zero filtering, all text preserved
        chunks = chunk_text_smartly(text)

        for chunk_idx, chunk in enumerate(chunks):
            chunk_prompt = prompt_instructions
            if len(chunks) > 1:
                chunk_prompt += (
                    f"\n\n[NOTE: Chunk {chunk_idx + 1} of {len(chunks)}. "
                    "Extract every record visible in this section independently.]"
                )
            records, tokens, cost = call_llm_for_extraction(
                chunk_prompt, chunk, api_key, model
            )
            all_records.extend(records)
            total_tokens += tokens
            total_cost += cost

    # Deduplicate records that appeared in overlapping chunk windows
    if len(all_records) > 1:
        all_records = deduplicate_records(all_records, col_names)

    # Fallback LLM sweep for fields that are universally N/A
    all_records, fb_tokens, fb_cost = _fallback_targeted_extraction(
        all_records, col_names, columns, text, images, api_key, model
    )
    total_tokens += fb_tokens
    total_cost += fb_cost

    if not all_records:
        row = {c: "No Data Found or Error" for c in col_names}
        row['Source File'] = filename
        return {'rows': [row], 'total_tokens': total_tokens, 'total_cost': total_cost}

    results = []
    for item in all_records:
        clean = {c: item.get(c, "N/A") for c in col_names}
        clean['Source File'] = filename
        results.append(clean)

    return {'rows': results, 'total_tokens': total_tokens, 'total_cost': total_cost}


# ---------------------------------------------------------------------------
# Batch processor
# ---------------------------------------------------------------------------

def process_pdfs_to_excel(pdf_paths: list, columns: list) -> tuple:
    """
    Process all PDFs in parallel and write a formatted Excel workbook.
    Returns (output_path: str, total_tokens: int, total_cost: float).
    """
    api_key = os.environ.get("OPENAI_API_KEY")
    learning_rules = get_learning_rules()

    max_workers = min(len(pdf_paths), 10)
    results_by_index: dict = {}

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_index = {
            executor.submit(process_single_pdf, path, columns, api_key, learning_rules): i
            for i, path in enumerate(pdf_paths)
        }
        for future in as_completed(future_to_index):
            idx = future_to_index[future]
            try:
                results_by_index[idx] = future.result()
            except Exception as e:
                filename = os.path.basename(pdf_paths[idx])
                col_names = _get_col_names(columns)
                row = {c: f"Error: {str(e)}" for c in col_names}
                row['Source File'] = filename
                results_by_index[idx] = {'rows': [row], 'total_tokens': 0, 'total_cost': 0.0}

    all_rows: list = []
    grand_total_tokens = 0
    grand_total_cost = 0.0

    for i in sorted(results_by_index.keys()):
        result = results_by_index[i]
        all_rows.extend(result['rows'])
        grand_total_tokens += result['total_tokens']
        grand_total_cost += result['total_cost']

    df = pd.DataFrame(all_rows)

    # Source File → first column
    cols = df.columns.tolist()
    if 'Source File' in cols:
        cols.insert(0, cols.pop(cols.index('Source File')))
        df = df[cols]

    output_filename = (
        os.path.basename(pdf_paths[0]) + "_extracted.xlsx"
        if len(pdf_paths) == 1
        else "US_Neuro_Batch_Extraction.xlsx"
    )
    output_path = f"/tmp/{output_filename}"

    # Write with full openpyxl formatting
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Extracted Data')
        ws = writer.sheets['Extracted Data']

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="0F172A")
        thin = Side(border_style="thin", color="D1D5DB")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Style header
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = cell_border
        ws.row_dimensions[1].height = 30

        # Auto-size columns (max 55 chars) and style data cells
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=0,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 55)
            for cell in col_cells[1:]:
                cell.border = cell_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.freeze_panes = "A2"

    return output_path, grand_total_tokens, grand_total_cost
